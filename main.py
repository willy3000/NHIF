import xlrd
from xlwt import Workbook
from difflib import SequenceMatcher
from pathlib import Path
import openpyxl
import os
import tkinter
from tkinter import *
from tkinter import messagebox
import tkinter as tk
from tkinter import scrolledtext
from ttkthemes import ThemedTk
from tkinter import ttk
from functools import partial
from tkinter import filedialog
import glob
import os
import time
import win32com.client
import threading


def browseFiles():
    filename = filedialog.askopenfilename(initialdir="/",
                                          title="Select a File",
                                          filetypes=(("xlsx files",
                                                      "*.xlsx*"),
                                                     ("xls files",
                                                      "*.xls*")))

    file.config(state = "normal")
    file.delete(0, END)
    file.insert(0,filename)
    file.config(foreground = "#000000")
    file.config(state="disable")


def success(holder):
    holder = holder
    success = Label(root, text=holder, padx=200, font=("helvetica", 12))
    success.config(bg="#3b9400", fg="#ffffff")
    success.place(x=20, y=450)
    time.sleep(3)
    success.destroy()

def success_thread(holder):
    threading.Thread(target=success, args=(holder,), daemon=True).start()

def failed(holder):
    holder=holder
    success = Label(root, text = holder,padx = 200,font = ("helvetica", 12))
    success.config(bg = "#eb1515",fg = "#ffffff")
    success.place(x = 10, y = 450)
    time.sleep(3)
    success.destroy()
def failed_thread(holder):
    threading.Thread(target=failed, args=(holder,), daemon=True).start()



def enterbutton(name,border):
    name["bg"] = "#f0f0f0"
    name["fg"] = "#212121"

    border["highlightbackground"] = "#242424"
    #b1["font"] = ("helvetica",11)
def leavebutton(name,border):
    name["bg"] = "#1d293d"
    name["fg"] = "#d1d1d1"
    name["font"] = ("helvetica", 10)
    border["highlightbackground"] = "#dbdbdb"



def validate_choices():
    global column_count
    if len(file.get()) > 0:
        path = file.get()
        wb = openpyxl.load_workbook(path, enumerate)
        # the 2 lines under do the same.
        sheet = wb.worksheets[0]
        row_count = sheet.max_row
        column_count = sheet.max_column

        stage = 0

        if int(rows.get()) > row_count:
            failed_thread("rows exceed maximum")
        else:
            stage+=1
        if int(column1.get()) > column_count:
            failed_thread("column exceeds maximun")
        else:
            stage+=1
        if int(column2.get()) > column_count:
            failed_thread("column exceeds maximun")
        else:
            stage+=1
        if int(column1.get()) == int(column2.get()):
            failed_thread("columns cannot be equal")
        else:
            stage+=1

        if stage == 4:
            sort()






def sort():
    path = file.get()
    path1 = file.get().split("/")
    infile = path1[len(path1)-1]
    name = infile.split(".")[0]
    file2 = infile

    xlApp = win32com.client.Dispatch('Excel.Application')

    xlWb = xlApp.Workbooks.Open(os.path.join(os.getcwd(), path))
    xlWb.SaveAs(os.path.join(os.getcwd(), "converted\\"+file2.split('.xlsx')[0] +
                             '.xls'), FileFormat=1)

    xlApp.Quit()
    outfile = name+".xls"

    wb = xlrd.open_workbook("converted\\"+outfile)
    wb2 = Workbook()
    sheet = wb.sheet_by_index(0)
    sheet1 = wb2.add_sheet('Sheet 1')


    row = 1
    county = []
    nhif = []
    count = 0
    num = 0

    range1 = int(rows.get())

    for x in range(range1-1):
        county_name = sheet.cell_value(row, int(column1.get()))
        nhif_name = sheet.cell_value(row, int(column2.get()))
        county_name = county_name.lower()
        nhif_name = nhif_name.lower()
        #ratio = SequenceMatcher(None,county_name,nhif_name).ratio()
        county = county_name.split(" ")
        print(county)
        nhif = nhif_name.split(" ")
        for y in nhif:
            for z in county:
                if y in z or z in y:
                    print(len(y))
                    count+=1

        if count <1:
            col = 0
            for x in range(column_count):
                sheet1.write(num, col, sheet.cell_value(row, col))
                col+=1
            num+=1
        row+=1
        count=0

    wb2.save("sorted\\"+name+'_sorted.xls')
    success_thread("operation successful")


root = ThemedTk(theme='adapta')
root.title("Duplicates")
root.config(bg="#1d293d")
root.geometry("500x500+500+100")
root.iconbitmap("assets\\nhif.ico")
root.resizable(False,False)
img = PhotoImage(file="assets\\logo.png")
Label(root, image=img).place(x=170, y=10)


filelabel = ttk.Label(root, text="File ", background="#1d293d", foreground="#d1d1d1")
filelabel.place(x=50, y=105)
file = ttk.Entry(root)
file.config(width=50)
file.config(state = "disable")
file.place(x=90, y=100)

button_border = tk.Frame(root, highlightbackground="#dbdbdb",
                          highlightthickness=0.5, bd=0)

choose = Button(button_border, text="Choose", font=("helvetica", 10), bg="#1d293d", fg="#d1d1d1",
             activebackground="#6c769e", borderwidth=0, command = browseFiles)
choose.bind("<Enter>", lambda event, btnname=choose,btnborder = button_border : enterbutton(btnname,btnborder))
choose.bind("<Leave>", lambda event, btnname=choose,btnborder = button_border : leavebutton(btnname,btnborder))
choose.pack()
button_border.place(x = 420, y = 100)

labelframe1 = LabelFrame(root, bg="#1d293d")
labelframe1.config(padx=20, pady=20)
labelframe1.place(x=50, y=170)

rowslabel = ttk.Label(labelframe1, text="No. of rows : ", background="#1d293d", foreground="#d1d1d1")
rowslabel.grid(row = 0, column = 0)
rows = ttk.Entry(labelframe1)
rows.config(width=30)
rows.grid(row = 0, column = 1)

comparelabel = ttk.Label(labelframe1, text="Compare : ", background="#1d293d", foreground="#d1d1d1")
comparelabel.grid(row = 1, column = 0,pady = 10)

frame1 = Frame(labelframe1, bg="#1d293d")
frame1.config(padx=100, pady=20)
frame1.grid(row=1, column=1)

column1 = ttk.Entry(frame1)
column1.config(width=5)
column1.grid(row = 0, column = 0)

tolabel = ttk.Label(frame1, text=" : to : ", background="#1d293d", foreground="#d1d1d1")
tolabel.grid(row = 0, column = 1)

column2 = ttk.Entry(frame1)
column2.config(width=5)
column2.grid(row = 0, column = 2)


button_border1 = tk.Frame(root, highlightbackground="#dbdbdb",
                          highlightthickness=0.5, bd=0)

process = Button(button_border1, text="Sort", font=("helvetica", 10), bg="#1d293d", fg="#d1d1d1",
             activebackground="#6c769e", borderwidth=0, command = validate_choices)
process.bind("<Enter>", lambda event, btnname=process,btnborder = button_border1 : enterbutton(btnname,btnborder))
process.bind("<Leave>", lambda event, btnname=process,btnborder = button_border1 : leavebutton(btnname,btnborder))
process.config(width = 30,height = 5)
process.pack()
button_border1.place(x = 150, y = 320)


root.mainloop()




